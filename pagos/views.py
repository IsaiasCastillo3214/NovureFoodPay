from datetime import timedelta
from functools import wraps

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import (
    Q,
    Sum,
    Count,
    Value,
    IntegerField,
    Prefetch,
    F,
    ExpressionWrapper,
)
from django.db.models.deletion import ProtectedError
from django.db.models.functions import Coalesce
from django.http import Http404, JsonResponse, HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.utils import timezone
from django.utils.dateparse import parse_date

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

from .forms import (
    PedidoForm,
    PedidoOwnerForm,
    ProductoForm,
    VendedorUsuarioForm,
)

from .models import (
    Pedido,
    Producto,
    DetallePedido,
    Vendedor,
)


# ============================================================
# ROLES Y PERMISOS
# ============================================================

def es_dueno(user):
    return user.is_superuser or user.groups.filter(name='Dueño local').exists()


def dueno_required(view_func):
    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect('login')

        if not es_dueno(request.user):
            messages.error(
                request,
                'No tienes permisos para acceder al panel del dueño.'
            )
            return redirect('dashboard')

        return view_func(request, *args, **kwargs)

    return wrapper


def obtener_vendedor_usuario(user):
    try:
        return user.perfil_vendedor
    except Vendedor.DoesNotExist:
        return None


def obtener_pedido_permitido(request, pedido_id):
    if es_dueno(request.user):
        return get_object_or_404(Pedido, id=pedido_id)

    vendedor = obtener_vendedor_usuario(request.user)

    if not vendedor:
        raise Http404('No tienes vendedor asociado.')

    return get_object_or_404(
        Pedido,
        id=pedido_id,
        vendedor=vendedor
    )


def construir_querystring(request, **overrides):
    query = request.GET.copy()

    if 'page' in query:
        query.pop('page')

    for key, value in overrides.items():
        if value is None or value == '':
            query.pop(key, None)
        elif key == 'tipo_entrega' and value == 'todos':
            query.pop(key, None)
        else:
            query[key] = value

    return query.urlencode()


# ============================================================
# FILTROS
# ============================================================

def aplicar_filtro_fecha(request, pedidos):
    rango_fecha = request.GET.get('rango_fecha', 'todos')
    fecha_inicio = request.GET.get('fecha_inicio', '').strip()
    fecha_fin = request.GET.get('fecha_fin', '').strip()

    hoy = timezone.localdate()

    if rango_fecha == 'hoy':
        pedidos = pedidos.filter(fecha__date=hoy)

    elif rango_fecha == 'semana':
        inicio = hoy - timedelta(days=7)
        pedidos = pedidos.filter(fecha__date__gte=inicio)

    elif rango_fecha == 'mes':
        inicio = hoy.replace(day=1)
        pedidos = pedidos.filter(fecha__date__gte=inicio)

    elif rango_fecha == 'personalizado':
        fecha_inicio_parseada = parse_date(fecha_inicio) if fecha_inicio else None
        fecha_fin_parseada = parse_date(fecha_fin) if fecha_fin else None

        if fecha_inicio_parseada:
            pedidos = pedidos.filter(fecha__date__gte=fecha_inicio_parseada)

        if fecha_fin_parseada:
            pedidos = pedidos.filter(fecha__date__lte=fecha_fin_parseada)

    return pedidos


def aplicar_busqueda(request, pedidos):
    busqueda = request.GET.get('q', '').strip()

    if not busqueda:
        return pedidos

    filtros_busqueda = (
        Q(nombre_cliente__icontains=busqueda) |
        Q(telefono_cliente__icontains=busqueda) |
        Q(detalle_entrega__icontains=busqueda) |
        Q(pedido__icontains=busqueda) |
        Q(vendedor__nombre__icontains=busqueda) |
        Q(detalles__producto__nombre__icontains=busqueda)
    )

    busqueda_numerica = ''.join(c for c in busqueda if c.isdigit())

    if busqueda_numerica:
        valor_numerico = int(busqueda_numerica)

        filtros_busqueda = (
            filtros_busqueda |
            Q(id=valor_numerico) |
            Q(monto=valor_numerico)
        )

    return pedidos.filter(filtros_busqueda).distinct()


def aplicar_filtro_tipo_entrega(request, pedidos):
    filtro_tipo_entrega = request.GET.get('tipo_entrega', 'todos')

    if filtro_tipo_entrega == 'delivery':
        return pedidos.filter(tipo_entrega='delivery')

    if filtro_tipo_entrega == 'retiro_tienda':
        return pedidos.filter(tipo_entrega='retiro_tienda')

    return pedidos


def obtener_pedidos_base_usuario(request):
    if es_dueno(request.user):
        return Pedido.objects.all()

    vendedor = obtener_vendedor_usuario(request.user)

    if not vendedor:
        messages.error(
            request,
            'Tu usuario no tiene un vendedor asociado.'
        )
        return Pedido.objects.none()

    return Pedido.objects.filter(vendedor=vendedor)


# ============================================================
# PRODUCTOS
# ============================================================

def obtener_productos_json(productos):
    return [
        {
            'id': str(producto.id),
            'nombre': producto.nombre,
            'precio': producto.precio,
        }
        for producto in productos
    ]


def obtener_lineas_productos(request):
    productos_ids = request.POST.getlist('producto[]')
    cantidades = request.POST.getlist('cantidad[]')

    lineas = []
    total = 0

    for producto_id, cantidad in zip(productos_ids, cantidades):
        if not producto_id:
            continue

        try:
            cantidad = int(cantidad)
        except ValueError:
            return None, 'La cantidad debe ser un número válido.'

        if cantidad <= 0:
            return None, 'La cantidad debe ser mayor a 0.'

        producto = Producto.objects.filter(
            id=producto_id,
            activo=True
        ).first()

        if not producto:
            return None, 'Uno de los productos seleccionados no existe o está inactivo.'

        precio_unitario = producto.precio
        total += precio_unitario * cantidad

        lineas.append({
            'producto': producto,
            'cantidad': cantidad,
            'precio_unitario': precio_unitario,
        })

    if not lineas:
        return None, 'Debes seleccionar al menos un producto.'

    return {
        'lineas': lineas,
        'total': total,
    }, None


def guardar_detalles_pedido(pedido, lineas):
    pedido.detalles.all().delete()

    total = 0

    for linea in lineas:
        DetallePedido.objects.create(
            pedido=pedido,
            producto=linea['producto'],
            cantidad=linea['cantidad'],
            precio_unitario=linea['precio_unitario'],
        )

        total += linea['cantidad'] * linea['precio_unitario']

    pedido.monto = total
    pedido.save(update_fields=['monto'])


# ============================================================
# DASHBOARD PRINCIPAL OPTIMIZADO
# ============================================================

@login_required
def dashboard(request):
    filtro_tipo_entrega = request.GET.get('tipo_entrega', 'todos')
    busqueda = request.GET.get('q', '').strip()
    rango_fecha = request.GET.get('rango_fecha', 'todos')
    fecha_inicio = request.GET.get('fecha_inicio', '').strip()
    fecha_fin = request.GET.get('fecha_fin', '').strip()
    pagina = request.GET.get('page', 1)

    pedidos_base = obtener_pedidos_base_usuario(request)

    pedidos_base = aplicar_filtro_fecha(request, pedidos_base)
    pedidos_base = aplicar_busqueda(request, pedidos_base)

    resumen = pedidos_base.aggregate(
        total_vendido=Coalesce(
            Sum('monto'),
            Value(0),
            output_field=IntegerField()
        ),
        total_pagado=Coalesce(
            Sum('monto', filter=Q(estado_pago='pagado')),
            Value(0),
            output_field=IntegerField()
        ),
        total_pendiente=Coalesce(
            Sum('monto', filter=Q(estado_pago='pendiente')),
            Value(0),
            output_field=IntegerField()
        ),
        total_delivery=Coalesce(
            Sum('monto', filter=Q(tipo_entrega='delivery')),
            Value(0),
            output_field=IntegerField()
        ),
        total_retiro=Coalesce(
            Sum('monto', filter=Q(tipo_entrega='retiro_tienda')),
            Value(0),
            output_field=IntegerField()
        ),
        cantidad_total=Count('id', distinct=True),
        cantidad_delivery=Count(
            'id',
            filter=Q(tipo_entrega='delivery'),
            distinct=True
        ),
        cantidad_retiro=Count(
            'id',
            filter=Q(tipo_entrega='retiro_tienda'),
            distinct=True
        ),
    )

    pedidos_filtrados = aplicar_filtro_tipo_entrega(request, pedidos_base)

    detalles_optimizados = Prefetch(
        'detalles',
        queryset=DetallePedido.objects.select_related('producto')
    )

    pedidos_filtrados = (
        pedidos_filtrados
        .select_related('vendedor')
        .prefetch_related(detalles_optimizados)
        .order_by('-fecha')
    )

    paginator = Paginator(pedidos_filtrados, 12)
    pedidos = paginator.get_page(pagina)

    query_sin_page = construir_querystring(request)
    query_sin_busqueda = construir_querystring(request, q=None)
    query_todos = construir_querystring(request, tipo_entrega='todos')
    query_delivery = construir_querystring(request, tipo_entrega='delivery')
    query_retiro = construir_querystring(request, tipo_entrega='retiro_tienda')

    context = {
        'pedidos': pedidos,

        'total_vendido': resumen['total_vendido'],
        'total_pagado': resumen['total_pagado'],
        'total_pendiente': resumen['total_pendiente'],

        'cantidad_total': resumen['cantidad_total'],
        'cantidad_delivery': resumen['cantidad_delivery'],
        'cantidad_retiro': resumen['cantidad_retiro'],
        'cantidad_pedidos_vista': pedidos.paginator.count,

        'total_delivery': resumen['total_delivery'],
        'total_retiro': resumen['total_retiro'],

        'filtro_tipo_entrega': filtro_tipo_entrega,
        'busqueda': busqueda,
        'rango_fecha': rango_fecha,
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin,

        'query_sin_page': query_sin_page,
        'query_sin_busqueda': query_sin_busqueda,
        'query_todos': query_todos,
        'query_delivery': query_delivery,
        'query_retiro': query_retiro,
        'export_querystring': query_sin_page,

        'estados_pedido': Pedido.ESTADO_PEDIDO_CHOICES,
        'estados_pago': Pedido.ESTADO_PAGO_CHOICES,
        'es_dueno': es_dueno(request.user),

        'page_obj': pedidos,
    }

    return render(request, 'pagos/dashboard.html', context)


# ============================================================
# EXPORTAR PEDIDOS A EXCEL
# ============================================================

@login_required
def exportar_pedidos_excel(request):
    pedidos = obtener_pedidos_base_usuario(request)

    pedidos = aplicar_filtro_fecha(request, pedidos)
    pedidos = aplicar_busqueda(request, pedidos)
    pedidos = aplicar_filtro_tipo_entrega(request, pedidos)

    pedidos = (
        pedidos
        .select_related('vendedor')
        .prefetch_related(
            Prefetch(
                'detalles',
                queryset=DetallePedido.objects.select_related('producto')
            )
        )
        .order_by('-fecha')
    )

    wb = Workbook()
    ws = wb.active
    ws.title = 'Pedidos'

    headers = [
        'ID Pedido',
        'Fecha',
        'Cliente',
        'Teléfono',
        'Vendedor',
        'Tipo entrega',
        'Detalle entrega',
        'Productos',
        'Monto',
        'Tipo pago',
        'Estado pedido',
        'Estado pago',
        'Observación',
    ]

    ws.append(headers)

    header_fill = PatternFill(
        start_color='0F172A',
        end_color='0F172A',
        fill_type='solid'
    )

    header_font = Font(
        color='FFFFFF',
        bold=True
    )

    border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0'),
    )

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for pedido in pedidos:
        productos_texto = []

        for detalle in pedido.detalles.all():
            productos_texto.append(
                f'{detalle.cantidad} x {detalle.producto.nombre} '
                f'(${detalle.subtotal})'
            )

        fecha_local = timezone.localtime(pedido.fecha).strftime('%d/%m/%Y %H:%M')

        ws.append([
            pedido.numero_pedido,
            fecha_local,
            pedido.nombre_cliente,
            pedido.telefono_cliente or '',
            pedido.vendedor.nombre,
            pedido.get_tipo_entrega_display(),
            pedido.detalle_entrega or '',
            '\n'.join(productos_texto),
            pedido.monto,
            pedido.get_tipo_pago_display(),
            pedido.get_estado_pedido_display(),
            pedido.get_estado_pago_display(),
            pedido.pedido or '',
        ])

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(
                vertical='top',
                wrap_text=True
            )

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions

    for column_cells in ws.columns:
        max_length = 0
        column = column_cells[0].column
        column_letter = get_column_letter(column)

        for cell in column_cells:
            value = str(cell.value) if cell.value is not None else ''

            if len(value) > max_length:
                max_length = len(value)

        adjusted_width = min(max_length + 3, 42)
        ws.column_dimensions[column_letter].width = adjusted_width

    ws.column_dimensions['H'].width = 38
    ws.column_dimensions['M'].width = 34

    for row_number in range(2, ws.max_row + 1):
        ws.row_dimensions[row_number].height = 42

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

    nombre_archivo = timezone.localtime().strftime('pedidos_%Y%m%d_%H%M.xlsx')

    response['Content-Disposition'] = (
        f'attachment; filename="{nombre_archivo}"'
    )

    wb.save(response)

    return response


# ============================================================
# CREAR PEDIDO
# ============================================================

@login_required
def crear_pedido(request):
    producto_error = None

    form_class = PedidoOwnerForm if es_dueno(request.user) else PedidoForm

    if request.method == 'POST':
        form = form_class(request.POST)
        resultado_productos, producto_error = obtener_lineas_productos(request)

        if form.is_valid() and not producto_error:
            with transaction.atomic():
                pedido = form.save(commit=False)

                if not es_dueno(request.user):
                    vendedor = obtener_vendedor_usuario(request.user)

                    if not vendedor:
                        messages.error(
                            request,
                            'Tu usuario no tiene un vendedor asociado.'
                        )
                        return redirect('dashboard')

                    pedido.vendedor = vendedor

                pedido.monto = resultado_productos['total']
                pedido.save()

                guardar_detalles_pedido(
                    pedido,
                    resultado_productos['lineas']
                )

            messages.success(request, 'Pedido registrado correctamente.')
            return redirect('dashboard')

    else:
        form = form_class()

    context = {
        'form': form,
        'producto_error': producto_error,
        'titulo': 'Registrar pedido',
        'boton': 'Guardar pedido',
        'detalles_existentes': None,
        'es_dueno': es_dueno(request.user),
    }

    return render(request, 'pagos/crear_pedido.html', context)


# ============================================================
# EDITAR PEDIDO
# ============================================================

@login_required
def editar_pedido(request, pedido_id):
    pedido = obtener_pedido_permitido(request, pedido_id)

    producto_error = None

    form_class = PedidoOwnerForm if es_dueno(request.user) else PedidoForm

    if request.method == 'POST':
        form = form_class(request.POST, instance=pedido)
        resultado_productos, producto_error = obtener_lineas_productos(request)

        if form.is_valid() and not producto_error:
            with transaction.atomic():
                pedido = form.save(commit=False)

                if not es_dueno(request.user):
                    vendedor = obtener_vendedor_usuario(request.user)

                    if not vendedor:
                        messages.error(
                            request,
                            'Tu usuario no tiene un vendedor asociado.'
                        )
                        return redirect('dashboard')

                    pedido.vendedor = vendedor

                pedido.monto = resultado_productos['total']
                pedido.save()

                guardar_detalles_pedido(
                    pedido,
                    resultado_productos['lineas']
                )

            messages.success(request, 'Pedido actualizado correctamente.')
            return redirect('dashboard')

    else:
        form = form_class(instance=pedido)

    context = {
        'form': form,
        'producto_error': producto_error,
        'titulo': 'Editar pedido',
        'boton': 'Actualizar pedido',
        'detalles_existentes': pedido.detalles.select_related('producto').all(),
        'es_dueno': es_dueno(request.user),
    }

    return render(request, 'pagos/crear_pedido.html', context)


# ============================================================
# ELIMINAR PEDIDO
# ============================================================

@login_required
def eliminar_pedido(request, pedido_id):
    pedido = obtener_pedido_permitido(request, pedido_id)

    if request.method == 'POST':
        pedido.delete()
        messages.success(request, 'Pedido eliminado correctamente.')
        return redirect('dashboard')

    context = {
        'pedido': pedido,
    }

    return render(request, 'pagos/confirmar_eliminar.html', context)


# ============================================================
# ACTUALIZAR ESTADO DEL PEDIDO Y PAGO
# ============================================================

@login_required
def actualizar_estado(request, pedido_id):
    pedido = obtener_pedido_permitido(request, pedido_id)

    if request.method == 'POST':
        nuevo_estado_pedido = request.POST.get('estado_pedido')
        nuevo_estado_pago = request.POST.get('estado_pago')

        estados_pedido_validos = [
            estado[0] for estado in Pedido.ESTADO_PEDIDO_CHOICES
        ]

        estados_pago_validos = [
            estado[0] for estado in Pedido.ESTADO_PAGO_CHOICES
        ]

        if nuevo_estado_pedido in estados_pedido_validos:
            pedido.estado_pedido = nuevo_estado_pedido

        if nuevo_estado_pago in estados_pago_validos:
            pedido.estado_pago = nuevo_estado_pago

        pedido.save()
        messages.success(request, 'Estado actualizado correctamente.')

    return redirect('dashboard')


# ============================================================
# BÚSQUEDA AJAX DE PRODUCTOS
# ============================================================

@login_required
def buscar_productos(request):
    busqueda = request.GET.get('q', '').strip()

    productos = Producto.objects.filter(activo=True)

    if busqueda:
        productos = productos.filter(nombre__icontains=busqueda)
    else:
        productos = productos.none()

    productos = productos.order_by('nombre').values(
        'id',
        'nombre',
        'precio'
    )[:20]

    return JsonResponse({
        'productos': list(productos)
    })


# ============================================================
# PANEL DEL DUEÑO - DASHBOARD
# ============================================================

@dueno_required
def panel_dueno(request):
    filtro_fecha = request.GET.get('rango', 'mes')

    pedidos = Pedido.objects.select_related('vendedor').all()

    hoy = timezone.localdate()

    if filtro_fecha == 'hoy':
        pedidos = pedidos.filter(fecha__date=hoy)

    elif filtro_fecha == 'semana':
        inicio = hoy - timedelta(days=7)
        pedidos = pedidos.filter(fecha__date__gte=inicio)

    elif filtro_fecha == 'mes':
        inicio = hoy.replace(day=1)
        pedidos = pedidos.filter(fecha__date__gte=inicio)

    elif filtro_fecha == 'todos':
        pedidos = pedidos.all()

    resumen = pedidos.aggregate(
        total_vendido=Coalesce(
            Sum('monto'),
            Value(0),
            output_field=IntegerField()
        ),
        total_pagado=Coalesce(
            Sum('monto', filter=Q(estado_pago='pagado')),
            Value(0),
            output_field=IntegerField()
        ),
        total_pendiente=Coalesce(
            Sum('monto', filter=Q(estado_pago='pendiente')),
            Value(0),
            output_field=IntegerField()
        ),
        total_pedidos=Count('id'),
        pedidos_delivery=Count(
            'id',
            filter=Q(tipo_entrega='delivery')
        ),
        pedidos_retiro=Count(
            'id',
            filter=Q(tipo_entrega='retiro_tienda')
        ),
        pedidos_pagados=Count(
            'id',
            filter=Q(estado_pago='pagado')
        ),
        pedidos_pendientes=Count(
            'id',
            filter=Q(estado_pago='pendiente')
        ),
    )

    total_pedidos = resumen['total_pedidos'] or 0
    total_vendido = resumen['total_vendido'] or 0

    if total_pedidos > 0:
        ticket_promedio = round(total_vendido / total_pedidos)
    else:
        ticket_promedio = 0

    subtotal_expression = ExpressionWrapper(
        F('cantidad') * F('precio_unitario'),
        output_field=IntegerField()
    )

    productos_top = (
        DetallePedido.objects
        .filter(pedido__in=pedidos)
        .values('producto__nombre')
        .annotate(
            cantidad_vendida=Coalesce(
                Sum('cantidad'),
                Value(0),
                output_field=IntegerField()
            ),
            total_vendido=Coalesce(
                Sum(subtotal_expression),
                Value(0),
                output_field=IntegerField()
            )
        )
        .order_by('-cantidad_vendida')[:6]
    )

    vendedores_top = (
        pedidos
        .values('vendedor__nombre')
        .annotate(
            total_vendido=Coalesce(
                Sum('monto'),
                Value(0),
                output_field=IntegerField()
            ),
            total_pedidos=Count('id')
        )
        .order_by('-total_vendido')[:6]
    )

    ultimos_pedidos = (
        pedidos
        .select_related('vendedor')
        .order_by('-fecha')[:8]
    )

    context = {
        'filtro_fecha': filtro_fecha,

        'total_vendido': total_vendido,
        'total_pagado': resumen['total_pagado'],
        'total_pendiente': resumen['total_pendiente'],
        'total_pedidos': total_pedidos,

        'pedidos_delivery': resumen['pedidos_delivery'],
        'pedidos_retiro': resumen['pedidos_retiro'],
        'pedidos_pagados': resumen['pedidos_pagados'],
        'pedidos_pendientes': resumen['pedidos_pendientes'],
        'ticket_promedio': ticket_promedio,

        'productos_top': productos_top,
        'vendedores_top': vendedores_top,
        'ultimos_pedidos': ultimos_pedidos,
    }

    return render(request, 'pagos/panel_dueno.html', context)


# ============================================================
# CRUD PRODUCTOS
# ============================================================

@dueno_required
def productos_lista(request):
    productos = Producto.objects.all().order_by('nombre')

    context = {
        'productos': productos,
    }

    return render(request, 'pagos/productos_lista.html', context)


@dueno_required
def producto_crear(request):
    if request.method == 'POST':
        form = ProductoForm(request.POST)

        if form.is_valid():
            form.save()
            messages.success(request, 'Producto creado correctamente.')
            return redirect('productos_lista')

    else:
        form = ProductoForm()

    context = {
        'form': form,
        'titulo': 'Crear producto',
        'boton': 'Guardar producto',
    }

    return render(request, 'pagos/producto_form.html', context)


@dueno_required
def producto_editar(request, producto_id):
    producto = get_object_or_404(Producto, id=producto_id)

    if request.method == 'POST':
        form = ProductoForm(request.POST, instance=producto)

        if form.is_valid():
            form.save()
            messages.success(request, 'Producto actualizado correctamente.')
            return redirect('productos_lista')

    else:
        form = ProductoForm(instance=producto)

    context = {
        'form': form,
        'titulo': 'Editar producto',
        'boton': 'Actualizar producto',
    }

    return render(request, 'pagos/producto_form.html', context)


@dueno_required
def producto_eliminar(request, producto_id):
    producto = get_object_or_404(Producto, id=producto_id)

    if request.method == 'POST':
        try:
            producto.delete()
            messages.success(
                request,
                'Producto eliminado correctamente.'
            )

        except ProtectedError:
            producto.activo = False
            producto.save(update_fields=['activo'])

            messages.warning(
                request,
                'Este producto ya tiene ventas asociadas. No se eliminó, pero fue marcado como inactivo.'
            )

        return redirect('productos_lista')

    context = {
        'producto': producto,
    }

    return render(
        request,
        'pagos/producto_confirmar_eliminar.html',
        context
    )


# ============================================================
# CRUD VENDEDORES Y CREDENCIALES
# ============================================================

@dueno_required
def vendedores_lista(request):
    vendedores = (
        Vendedor.objects
        .select_related('usuario')
        .all()
        .order_by('nombre')
    )

    context = {
        'vendedores': vendedores,
    }

    return render(request, 'pagos/vendedores_lista.html', context)


@dueno_required
def vendedor_crear(request):
    if request.method == 'POST':
        form = VendedorUsuarioForm(request.POST)

        if form.is_valid():
            form.save()
            messages.success(request, 'Vendedor creado correctamente.')
            return redirect('vendedores_lista')

    else:
        form = VendedorUsuarioForm()

    context = {
        'form': form,
        'titulo': 'Crear vendedor',
        'boton': 'Guardar vendedor',
        'editando': False,
    }

    return render(request, 'pagos/vendedor_form.html', context)


@dueno_required
def vendedor_editar(request, vendedor_id):
    vendedor = get_object_or_404(Vendedor, id=vendedor_id)

    if request.method == 'POST':
        form = VendedorUsuarioForm(
            request.POST,
            instance=vendedor
        )

        if form.is_valid():
            form.save()
            messages.success(request, 'Vendedor actualizado correctamente.')
            return redirect('vendedores_lista')

    else:
        form = VendedorUsuarioForm(instance=vendedor)

    context = {
        'form': form,
        'titulo': 'Editar vendedor',
        'boton': 'Actualizar vendedor',
        'editando': True,
    }

    return render(request, 'pagos/vendedor_form.html', context)


@dueno_required
def vendedor_eliminar(request, vendedor_id):
    vendedor = get_object_or_404(
        Vendedor.objects.select_related('usuario'),
        id=vendedor_id
    )

    if request.method == 'POST':
        usuario = vendedor.usuario

        tiene_pedidos = Pedido.objects.filter(vendedor=vendedor).exists()

        if tiene_pedidos:
            if usuario:
                usuario.is_active = False
                usuario.save(update_fields=['is_active'])

            messages.warning(
                request,
                'Este vendedor tiene pedidos asociados. No fue eliminado, pero su usuario fue desactivado.'
            )

            return redirect('vendedores_lista')

        vendedor.delete()

        if usuario:
            usuario.delete()

        messages.success(request, 'Vendedor eliminado correctamente.')
        return redirect('vendedores_lista')

    context = {
        'vendedor': vendedor,
    }

    return render(
        request,
        'pagos/vendedor_confirmar_eliminar.html',
        context
    )