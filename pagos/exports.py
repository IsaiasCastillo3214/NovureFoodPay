from django.contrib.auth.decorators import login_required
from django.db.models import Prefetch
from django.http import HttpResponse
from django.utils import timezone

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

from .models import DetallePedido
from .permissions import (
    negocio_activo_required,
    obtener_negocio_activo,
)
from .selectors import (
    aplicar_filtro_fecha,
    aplicar_busqueda,
    aplicar_filtro_tipo_entrega,
    obtener_pedidos_base_usuario,
)


# ============================================================
# EXPORTACIONES
# ============================================================

@login_required
@negocio_activo_required
def exportar_pedidos_excel(request):
    pedidos = obtener_pedidos_base_usuario(request)

    pedidos = aplicar_filtro_fecha(request, pedidos)
    pedidos = aplicar_busqueda(request, pedidos)
    pedidos = aplicar_filtro_tipo_entrega(request, pedidos)

    pedidos = (
        pedidos
        .select_related('vendedor', 'negocio')
        .prefetch_related(
            Prefetch(
                'detalles',
                queryset=DetallePedido.objects.select_related('producto')
            )
        )
        .order_by('-fecha')
    )

    wb = Workbook()
    ws = wb.active
    ws.title = 'Pedidos'

    headers = [
        'Negocio',
        'ID Pedido',
        'Fecha',
        'Cliente',
        'Teléfono',
        'Vendedor',
        'Tipo entrega',
        'Detalle entrega',
        'Productos',
        'Monto',
        'Tipo pago',
        'Estado pedido',
        'Estado pago',
        'Observación',
    ]

    ws.append(headers)

    header_fill = PatternFill(
        start_color='0F172A',
        end_color='0F172A',
        fill_type='solid'
    )

    header_font = Font(
        color='FFFFFF',
        bold=True
    )

    border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0'),
    )

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for pedido in pedidos:
        productos_texto = []

        for detalle in pedido.detalles.all():
            productos_texto.append(
                f'{detalle.cantidad} x {detalle.producto.nombre} '
                f'(${detalle.subtotal})'
            )

        fecha_local = timezone.localtime(pedido.fecha).strftime('%d/%m/%Y %H:%M')

        ws.append([
            pedido.negocio.nombre if pedido.negocio else '',
            pedido.numero_pedido,
            fecha_local,
            pedido.nombre_cliente,
            pedido.telefono_cliente or '',
            pedido.vendedor.nombre,
            pedido.get_tipo_entrega_display(),
            pedido.detalle_entrega or '',
            '\n'.join(productos_texto),
            pedido.monto,
            pedido.get_tipo_pago_display(),
            pedido.get_estado_pedido_display(),
            pedido.get_estado_pago_display(),
            pedido.pedido or '',
        ])

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(
                vertical='top',
                wrap_text=True
            )

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions

    for column_cells in ws.columns:
        max_length = 0
        column = column_cells[0].column
        column_letter = get_column_letter(column)

        for cell in column_cells:
            value = str(cell.value) if cell.value is not None else ''

            if len(value) > max_length:
                max_length = len(value)

        adjusted_width = min(max_length + 3, 42)
        ws.column_dimensions[column_letter].width = adjusted_width

    ws.column_dimensions['I'].width = 38
    ws.column_dimensions['N'].width = 34

    for row_number in range(2, ws.max_row + 1):
        ws.row_dimensions[row_number].height = 42

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

    negocio = obtener_negocio_activo(request)
    nombre_negocio = negocio.slug if negocio else 'pedidos'
    fecha_archivo = timezone.localtime().strftime('%Y%m%d_%H%M')
    nombre_archivo = f'pedidos_{nombre_negocio}_{fecha_archivo}.xlsx'

    response['Content-Disposition'] = (
        f'attachment; filename="{nombre_archivo}"'
    )

    wb.save(response)

    return response
